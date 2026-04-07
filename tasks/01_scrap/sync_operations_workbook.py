import hashlib
import os
import re
import sqlite3
import tempfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlparse

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DATA_DIR = PROJECT_ROOT / "data"
DATA_DIR = Path(os.environ.get("CRM_DATA_DIR") or DEFAULT_DATA_DIR)
OPERATIONS_DB_PATH = Path(
    os.environ.get("OPERATIONS_DB_PATH") or DATA_DIR / "dashboard_operations.db"
)
WORKBOOK_URL = (os.environ.get("OPERATIONS_WORKBOOK_URL") or "").strip()
BUILDER_VERSION = "operations-v1"
HTTP_TIMEOUT_SECONDS = 120
VALID_CUSTOMER_ID_PATTERN = re.compile(r"^KH\d+", re.IGNORECASE)


def log(message: str):
    print(f"[operations-sync] {message}")


def normalize_text(value):
    return str(value or "").strip()


def fold_text(value):
    text = normalize_text(value)
    replacements = {
        "đ": "d",
        "Đ": "d",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return (
        text.encode("ascii", "ignore").decode("ascii").lower().strip()
    )


def normalize_sheet_name(value):
    return re.sub(r"\s+", " ", fold_text(value))


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(value)).date()
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def format_date_key(value):
    parsed = parse_date(value)
    return parsed.isoformat() if parsed else None


def end_of_month(value):
    parsed = parse_date(value)
    if not parsed:
        return None
    if parsed.month == 12:
        return date(parsed.year, 12, 31)
    next_month = date(parsed.year, parsed.month + 1, 1)
    return next_month.fromordinal(next_month.toordinal() - 1)


def format_month_end_key(value):
    month_end = end_of_month(value)
    return month_end.isoformat() if month_end else None


def to_int(value):
    if value is None or value == "":
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def to_float(value):
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def to_bool_int(value):
    if value in (True, False):
        return 1 if value else 0
    if value is None or value == "":
        return 0
    try:
        return 1 if float(value) > 0 else 0
    except (TypeError, ValueError):
        return 0


def normalize_account(value):
    return normalize_text(value)


def normalize_customer_id(value):
    text = normalize_text(value)
    if not text:
        return None
    return text if VALID_CUSTOMER_ID_PATTERN.match(text) else None


def is_renew_contract(contract_term):
    return "RENEW" in normalize_text(contract_term).upper()


def get_sheet(workbook, preferred_name):
    expected = normalize_sheet_name(preferred_name)
    for sheet_name in workbook.sheetnames:
        if normalize_sheet_name(sheet_name) == expected:
            return workbook[sheet_name]
    raise KeyError(f"Worksheet not found: {preferred_name}")


def download_workbook(url):
    if not url:
        raise RuntimeError("Missing OPERATIONS_WORKBOOK_URL in environment.")

    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        raise RuntimeError("OPERATIONS_WORKBOOK_URL must be an http/https URL.")

    log("downloading workbook export")
    response = requests.get(url, timeout=HTTP_TIMEOUT_SECONDS)
    response.raise_for_status()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        tmp.write(response.content)
        tmp.flush()
        return Path(tmp.name)
    finally:
        tmp.close()


def parse_activation_sheet(workbook):
    sheet = get_sheet(workbook, "Activation")
    rows = []
    seen_accounts = set()

    for values in sheet.iter_rows(min_row=2, values_only=True):
        account = normalize_account(values[0] if len(values) > 0 else None)
        if not account or account in seen_accounts:
            continue

        seen_accounts.add(account)
        row = {
            "account": account,
            "customer_type": normalize_text(values[1] if len(values) > 1 else None) or None,
            "customer_id": normalize_customer_id(values[2] if len(values) > 2 else None),
            "customer_name": normalize_text(values[3] if len(values) > 3 else None) or None,
            "sale_owner": normalize_text(values[4] if len(values) > 4 else None) or None,
            "activation_date": format_date_key(values[5] if len(values) > 5 else None),
            "expiry_date": format_date_key(values[6] if len(values) > 6 else None),
            "contract_term": normalize_text(values[7] if len(values) > 7 else None) or None,
            "account_type": normalize_text(values[8] if len(values) > 8 else None) or None,
        }
        row["activation_month_end"] = format_month_end_key(values[5] if len(values) > 5 else None)
        row["expiry_month_end"] = format_month_end_key(values[6] if len(values) > 6 else None)
        row["is_renew_contract"] = 1 if is_renew_contract(row["contract_term"]) else 0
        rows.append(row)

    return rows


def parse_jcd_sheet(workbook):
    sheet = get_sheet(workbook, "JCD hết hạn")
    rows = []
    seen_keys = set()

    for values in sheet.iter_rows(min_row=2, values_only=True):
        account = normalize_account(values[0] if len(values) > 0 else None)
        expiry_date = format_date_key(values[2] if len(values) > 2 else None)
        key = (account, expiry_date)
        if not account or key in seen_keys:
            continue

        seen_keys.add(key)
        password_value = normalize_text(values[4] if len(values) > 4 else None) or None
        rows.append({
            "account": account,
            "activation_date": format_date_key(values[1] if len(values) > 1 else None),
            "expiry_date": expiry_date,
            "account_type": normalize_text(values[3] if len(values) > 3 else None) or None,
            "password_value": password_value,
            "password_customer_id": normalize_customer_id(password_value),
            "activation_month_end": format_month_end_key(values[1] if len(values) > 1 else None),
            "expiry_month_end": format_month_end_key(values[2] if len(values) > 2 else None),
        })

    return rows


def parse_raw_sheet(workbook):
    sheet = get_sheet(workbook, "Raw Data")
    rows = []
    latest_date = None
    unique_accounts = set()
    invalid_daily_rows = 0

    for values in sheet.iter_rows(min_row=2, values_only=True):
        day_key = format_date_key(values[0] if len(values) > 0 else None)
        account = normalize_account(values[1] if len(values) > 1 else None)
        if not day_key or not account:
            continue

        open_cnt = to_int(values[4] if len(values) > 4 else None)
        create_cnt = to_int(values[5] if len(values) > 5 else None)
        update_cnt = to_int(values[6] if len(values) > 6 else None)
        render_cnt = to_int(values[7] if len(values) > 7 else None)
        quality_flag = to_bool_int(values[12] if len(values) > 12 else None)
        open_flag = to_bool_int(values[13] if len(values) > 13 else None)
        invalid_daily = 1 if open_cnt == 0 and (create_cnt > 0 or update_cnt > 0 or render_cnt > 0) else 0
        if invalid_daily:
            invalid_daily_rows += 1

        rows.append({
            "day_key": day_key,
            "account": account,
            "regdate": format_date_key(values[2] if len(values) > 2 else None),
            "organization_name": normalize_text(values[3] if len(values) > 3 else None) or None,
            "open_cnt": open_cnt,
            "create_cnt": create_cnt,
            "update_cnt": update_cnt,
            "render_cnt": render_cnt,
            "month_end_key": format_month_end_key(values[11] if len(values) > 11 else None) or format_month_end_key(day_key),
            "quality_flag": quality_flag,
            "open_flag": open_flag,
            "invalid_daily": invalid_daily,
        })

        latest_date = max(latest_date, day_key) if latest_date else day_key
        unique_accounts.add(account)

    return {
        "rows": rows,
        "latest_date": latest_date,
        "unique_accounts": unique_accounts,
        "invalid_daily_rows": invalid_daily_rows,
    }


def parse_month_matrix(workbook, sheet_name):
    sheet = get_sheet(workbook, sheet_name)
    month_keys = []
    for cell in sheet[1][4:]:
        month_key = format_month_end_key(cell.value)
        if month_key:
            month_keys.append(month_key)

    result = defaultdict(dict)
    for values in sheet.iter_rows(min_row=2, values_only=True):
        account = normalize_account(values[0] if len(values) > 0 else None)
        if not account:
            continue
        for index, month_key in enumerate(month_keys, start=4):
            value = values[index] if len(values) > index else None
            result[account][month_key] = normalize_text(value) if isinstance(value, str) else value

    return dict(result), month_keys


def parse_definition_thresholds(workbook):
    sheet = get_sheet(workbook, "Definition")
    low_open_threshold = 1.0
    high_open_threshold = 13.0
    quality_threshold = 0.35

    for values in sheet.iter_rows(min_row=1, max_col=4, values_only=True):
        label = normalize_text(values[1] if len(values) > 1 else None).lower()
        open_value = values[2] if len(values) > 2 else None
        quality_value = values[3] if len(values) > 3 else None

        if label == "ghost":
            low_open_threshold = to_float(open_value)
        elif label == "noise":
            high_open_threshold = to_float(open_value)
        elif label == "best":
            quality_threshold = to_float(quality_value)

    return {
        "low_open_threshold": low_open_threshold,
        "high_open_threshold": high_open_threshold,
        "quality_threshold": quality_threshold,
    }


def build_monthly_metrics(raw_rows):
    monthly = {}
    latest_active_by_account = {}

    for row in raw_rows:
        account = row["account"]
        month_key = row["month_end_key"]
        key = (account, month_key)
        if key not in monthly:
            monthly[key] = {
                "account": account,
                "month_end_key": month_key,
                "open_cnt": 0,
                "create_cnt": 0,
                "update_cnt": 0,
                "render_cnt": 0,
                "quality_numerator": 0,
                "open_days": 0,
                "invalid_daily_count": 0,
                "latest_active_date": None,
            }

        target = monthly[key]
        target["open_cnt"] += row["open_cnt"]
        target["create_cnt"] += row["create_cnt"]
        target["update_cnt"] += row["update_cnt"]
        target["render_cnt"] += row["render_cnt"]
        target["quality_numerator"] += row["quality_flag"]
        target["open_days"] += row["open_flag"]
        target["invalid_daily_count"] += row["invalid_daily"]

        if row["open_cnt"] > 0 or row["create_cnt"] > 0 or row["update_cnt"] > 0 or row["render_cnt"] > 0:
            current_latest = target["latest_active_date"]
            target["latest_active_date"] = max(current_latest, row["day_key"]) if current_latest else row["day_key"]
            previous_latest = latest_active_by_account.get(account)
            latest_active_by_account[account] = max(previous_latest, row["day_key"]) if previous_latest else row["day_key"]

    for target in monthly.values():
        open_days = target["open_days"]
        target["quality_ratio"] = round((target["quality_numerator"] / open_days), 6) if open_days > 0 else 0.0

    return list(monthly.values()), latest_active_by_account


def build_monthly_statuses(activation_rows, status_map, category_map, month_keys):
    rows = []
    for activation in activation_rows:
        account = activation["account"]
        account_statuses = status_map.get(account, {})
        account_categories = category_map.get(account, {})
        for month_key in month_keys:
            rows.append({
                "account": account,
                "month_end_key": month_key,
                "status": normalize_text(account_statuses.get(month_key)) or None,
                "category": normalize_text(account_categories.get(month_key)) or None,
            })
    return rows


def build_due_accounts(activation_rows, jcd_rows):
    activation_by_account = {row["account"]: row for row in activation_rows}
    renew_rows_by_account = defaultdict(list)
    for row in activation_rows:
        if row["is_renew_contract"]:
            renew_rows_by_account[row["account"]].append(row)

    due_accounts = {}

    for row in jcd_rows:
        account = row["account"]
        due_month_key = row["expiry_month_end"]
        due_date = row["expiry_date"]
        if not account or not due_month_key or not due_date:
            continue

        activation = activation_by_account.get(account)
        renew_match = None
        for candidate in sorted(
            renew_rows_by_account.get(account, []),
            key=lambda item: item["activation_date"] or "",
        ):
            if candidate["activation_date"] and candidate["activation_date"] >= due_date:
                renew_match = candidate
                break

        due_accounts[(account, due_month_key)] = {
            "account": account,
            "due_month_key": due_month_key,
            "due_date": due_date,
            "source": "jcd",
            "customer_type": activation["customer_type"] if activation else None,
            "customer_id": activation["customer_id"] if activation else None,
            "customer_name": activation["customer_name"] if activation else None,
            "sale_owner": activation["sale_owner"] if activation else None,
            "account_type": activation["account_type"] if activation else row["account_type"],
            "renewed": 1 if renew_match else 0,
            "renew_activation_date": renew_match["activation_date"] if renew_match else None,
            "renew_expiry_date": renew_match["expiry_date"] if renew_match else None,
            "current_expiry_date": activation["expiry_date"] if activation else None,
        }

    for row in activation_rows:
        account = row["account"]
        due_month_key = row["expiry_month_end"]
        if not account or not due_month_key:
            continue

        key = (account, due_month_key)
        if key in due_accounts:
            continue

        due_accounts[key] = {
            "account": account,
            "due_month_key": due_month_key,
            "due_date": row["expiry_date"],
            "source": "activation",
            "customer_type": row["customer_type"],
            "customer_id": row["customer_id"],
            "customer_name": row["customer_name"],
            "sale_owner": row["sale_owner"],
            "account_type": row["account_type"],
            "renewed": 0,
            "renew_activation_date": None,
            "renew_expiry_date": None,
            "current_expiry_date": row["expiry_date"],
        }

    return list(due_accounts.values())


def build_meta_payload(raw_info, activation_rows, jcd_rows, due_accounts, latest_status_month, definition_thresholds):
    activation_accounts = {row["account"] for row in activation_rows}
    jcd_accounts = {row["account"] for row in jcd_rows}
    raw_accounts = raw_info["unique_accounts"]
    bridged_accounts = activation_accounts | jcd_accounts
    default_report_month = format_month_end_key(raw_info["latest_date"]) if raw_info["latest_date"] else latest_status_month
    items = {
        "builder_version": BUILDER_VERSION,
        "built_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "latest_raw_date": raw_info["latest_date"] or "",
        "default_report_month": default_report_month or "",
        "latest_status_month": latest_status_month or "",
        "raw_unique_accounts": str(len(raw_accounts)),
        "activation_unique_accounts": str(len(activation_accounts)),
        "jcd_unique_accounts": str(len(jcd_accounts)),
        "bridged_unique_accounts": str(len(bridged_accounts)),
        "raw_accounts_in_bridge": str(len(raw_accounts & bridged_accounts)),
        "raw_accounts_excluded": str(len(raw_accounts - bridged_accounts)),
        "invalid_daily_rows": str(raw_info["invalid_daily_rows"]),
        "due_accounts_total": str(len(due_accounts)),
        "workbook_url_sha1": hashlib.sha1(WORKBOOK_URL.encode("utf-8")).hexdigest() if WORKBOOK_URL else "",
        "threshold_open_low": str(definition_thresholds["low_open_threshold"]),
        "threshold_open_high": str(definition_thresholds["high_open_threshold"]),
        "threshold_quality": str(definition_thresholds["quality_threshold"]),
    }
    return items


def init_db(conn):
    conn.executescript(
        """
        PRAGMA journal_mode=WAL;
        PRAGMA busy_timeout=60000;

        CREATE TABLE IF NOT EXISTS operations_meta (
          key TEXT PRIMARY KEY,
          value TEXT
        );

        CREATE TABLE IF NOT EXISTS ops_activation_accounts (
          account TEXT PRIMARY KEY,
          customer_type TEXT,
          customer_id TEXT,
          customer_name TEXT,
          sale_owner TEXT,
          activation_date TEXT,
          activation_month_end TEXT,
          expiry_date TEXT,
          expiry_month_end TEXT,
          contract_term TEXT,
          account_type TEXT,
          is_renew_contract INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS ops_jcd_expired_accounts (
          account TEXT NOT NULL,
          activation_date TEXT,
          activation_month_end TEXT,
          expiry_date TEXT,
          expiry_month_end TEXT,
          account_type TEXT,
          password_value TEXT,
          password_customer_id TEXT,
          PRIMARY KEY (account, expiry_date)
        );

        CREATE TABLE IF NOT EXISTS ops_raw_daily (
          day_key TEXT NOT NULL,
          account TEXT NOT NULL,
          regdate TEXT,
          organization_name TEXT,
          open_cnt INTEGER NOT NULL DEFAULT 0,
          create_cnt INTEGER NOT NULL DEFAULT 0,
          update_cnt INTEGER NOT NULL DEFAULT 0,
          render_cnt INTEGER NOT NULL DEFAULT 0,
          month_end_key TEXT NOT NULL,
          quality_flag INTEGER NOT NULL DEFAULT 0,
          open_flag INTEGER NOT NULL DEFAULT 0,
          invalid_daily INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS ops_monthly_metrics (
          account TEXT NOT NULL,
          month_end_key TEXT NOT NULL,
          open_cnt INTEGER NOT NULL DEFAULT 0,
          create_cnt INTEGER NOT NULL DEFAULT 0,
          update_cnt INTEGER NOT NULL DEFAULT 0,
          render_cnt INTEGER NOT NULL DEFAULT 0,
          quality_numerator INTEGER NOT NULL DEFAULT 0,
          open_days INTEGER NOT NULL DEFAULT 0,
          quality_ratio REAL NOT NULL DEFAULT 0,
          invalid_daily_count INTEGER NOT NULL DEFAULT 0,
          latest_active_date TEXT,
          PRIMARY KEY (account, month_end_key)
        );

        CREATE TABLE IF NOT EXISTS ops_monthly_status (
          account TEXT NOT NULL,
          month_end_key TEXT NOT NULL,
          status TEXT,
          category TEXT,
          PRIMARY KEY (account, month_end_key)
        );

        CREATE TABLE IF NOT EXISTS ops_due_accounts (
          account TEXT NOT NULL,
          due_month_key TEXT NOT NULL,
          due_date TEXT,
          source TEXT NOT NULL,
          customer_type TEXT,
          customer_id TEXT,
          customer_name TEXT,
          sale_owner TEXT,
          account_type TEXT,
          renewed INTEGER NOT NULL DEFAULT 0,
          renew_activation_date TEXT,
          renew_expiry_date TEXT,
          current_expiry_date TEXT,
          PRIMARY KEY (account, due_month_key, source)
        );

        CREATE INDEX IF NOT EXISTS idx_ops_raw_daily_account ON ops_raw_daily(account);
        CREATE INDEX IF NOT EXISTS idx_ops_raw_daily_month ON ops_raw_daily(month_end_key);
        CREATE INDEX IF NOT EXISTS idx_ops_monthly_metrics_month ON ops_monthly_metrics(month_end_key);
        CREATE INDEX IF NOT EXISTS idx_ops_monthly_status_month ON ops_monthly_status(month_end_key);
        CREATE INDEX IF NOT EXISTS idx_ops_due_accounts_month ON ops_due_accounts(due_month_key);
        """
    )


def insert_many(conn, sql, rows):
    if not rows:
        return
    conn.executemany(sql, rows)


def write_database(
    activation_rows,
    jcd_rows,
    raw_rows,
    monthly_metrics,
    monthly_status_rows,
    due_accounts,
    meta_items,
):
    OPERATIONS_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    temp_db_path = OPERATIONS_DB_PATH.parent / f"{OPERATIONS_DB_PATH.name}.{os.getpid()}.{int(datetime.utcnow().timestamp())}.tmp"

    if temp_db_path.exists():
        temp_db_path.unlink()

    conn = sqlite3.connect(temp_db_path)
    try:
        init_db(conn)

        insert_many(
            conn,
            """
            INSERT INTO operations_meta(key, value)
            VALUES(?, ?)
            """,
            list(meta_items.items()),
        )
        insert_many(
            conn,
            """
            INSERT INTO ops_activation_accounts(
              account, customer_type, customer_id, customer_name, sale_owner,
              activation_date, activation_month_end, expiry_date, expiry_month_end,
              contract_term, account_type, is_renew_contract
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["account"],
                    row["customer_type"],
                    row["customer_id"],
                    row["customer_name"],
                    row["sale_owner"],
                    row["activation_date"],
                    row["activation_month_end"],
                    row["expiry_date"],
                    row["expiry_month_end"],
                    row["contract_term"],
                    row["account_type"],
                    row["is_renew_contract"],
                )
                for row in activation_rows
            ],
        )
        insert_many(
            conn,
            """
            INSERT INTO ops_jcd_expired_accounts(
              account, activation_date, activation_month_end, expiry_date,
              expiry_month_end, account_type, password_value, password_customer_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["account"],
                    row["activation_date"],
                    row["activation_month_end"],
                    row["expiry_date"],
                    row["expiry_month_end"],
                    row["account_type"],
                    row["password_value"],
                    row["password_customer_id"],
                )
                for row in jcd_rows
            ],
        )
        insert_many(
            conn,
            """
            INSERT INTO ops_raw_daily(
              day_key, account, regdate, organization_name, open_cnt, create_cnt,
              update_cnt, render_cnt, month_end_key, quality_flag, open_flag, invalid_daily
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["day_key"],
                    row["account"],
                    row["regdate"],
                    row["organization_name"],
                    row["open_cnt"],
                    row["create_cnt"],
                    row["update_cnt"],
                    row["render_cnt"],
                    row["month_end_key"],
                    row["quality_flag"],
                    row["open_flag"],
                    row["invalid_daily"],
                )
                for row in raw_rows
            ],
        )
        insert_many(
            conn,
            """
            INSERT INTO ops_monthly_metrics(
              account, month_end_key, open_cnt, create_cnt, update_cnt, render_cnt,
              quality_numerator, open_days, quality_ratio, invalid_daily_count, latest_active_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["account"],
                    row["month_end_key"],
                    row["open_cnt"],
                    row["create_cnt"],
                    row["update_cnt"],
                    row["render_cnt"],
                    row["quality_numerator"],
                    row["open_days"],
                    row["quality_ratio"],
                    row["invalid_daily_count"],
                    row["latest_active_date"],
                )
                for row in monthly_metrics
            ],
        )
        insert_many(
            conn,
            """
            INSERT INTO ops_monthly_status(account, month_end_key, status, category)
            VALUES (?, ?, ?, ?)
            """,
            [
                (
                    row["account"],
                    row["month_end_key"],
                    row["status"],
                    row["category"],
                )
                for row in monthly_status_rows
            ],
        )
        insert_many(
            conn,
            """
            INSERT INTO ops_due_accounts(
              account, due_month_key, due_date, source, customer_type, customer_id,
              customer_name, sale_owner, account_type, renewed, renew_activation_date,
              renew_expiry_date, current_expiry_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["account"],
                    row["due_month_key"],
                    row["due_date"],
                    row["source"],
                    row["customer_type"],
                    row["customer_id"],
                    row["customer_name"],
                    row["sale_owner"],
                    row["account_type"],
                    row["renewed"],
                    row["renew_activation_date"],
                    row["renew_expiry_date"],
                    row["current_expiry_date"],
                )
                for row in due_accounts
            ],
        )

        conn.commit()
    finally:
        conn.close()

    if OPERATIONS_DB_PATH.exists():
        OPERATIONS_DB_PATH.unlink()
    temp_db_path.rename(OPERATIONS_DB_PATH)


def main():
    workbook_path = None
    try:
        workbook_path = download_workbook(WORKBOOK_URL)
        workbook = load_workbook(workbook_path, data_only=True, read_only=False)

        activation_rows = parse_activation_sheet(workbook)
        jcd_rows = parse_jcd_sheet(workbook)
        raw_info = parse_raw_sheet(workbook)
        check_active_map, month_keys = parse_month_matrix(workbook, "Check_Active")
        check_categories_map, category_month_keys = parse_month_matrix(workbook, "Check_Categories")
        definition_thresholds = parse_definition_thresholds(workbook)
        unified_month_keys = sorted(set(month_keys) | set(category_month_keys))
        monthly_metrics, _latest_active_map = build_monthly_metrics(raw_info["rows"])
        monthly_status_rows = build_monthly_statuses(
            activation_rows,
            check_active_map,
            check_categories_map,
            unified_month_keys,
        )
        due_accounts = build_due_accounts(activation_rows, jcd_rows)
        meta_items = build_meta_payload(
            raw_info,
            activation_rows,
            jcd_rows,
            due_accounts,
            max(unified_month_keys) if unified_month_keys else None,
            definition_thresholds,
        )

        write_database(
            activation_rows=activation_rows,
            jcd_rows=jcd_rows,
            raw_rows=raw_info["rows"],
            monthly_metrics=monthly_metrics,
            monthly_status_rows=monthly_status_rows,
            due_accounts=due_accounts,
            meta_items=meta_items,
        )

        log(
            f"built operations db at {OPERATIONS_DB_PATH} "
            f"(activation={len(activation_rows)}, raw={len(raw_info['rows'])}, due={len(due_accounts)})"
        )
    finally:
        if workbook_path and workbook_path.exists():
            try:
                workbook_path.unlink()
            except OSError:
                pass


if __name__ == "__main__":
    main()
